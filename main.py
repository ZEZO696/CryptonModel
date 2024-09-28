import streamlit as st
import requests
import openpyxl
from datetime import datetime, timedelta
from sklearn.linear_model import LinearRegression
from statsmodels.tsa.arima.model import ARIMA
import itertools
import os

# Check if the 'reports' folder exists and create it if it doesn't
if not os.path.exists('reports'):
    os.mkdir('reports')

# Define the CryptoCompare API endpoint
CRYPTOCOMPARE_API_ENDPOINT = "https://min-api.cryptocompare.com/data/v2/histohour"

# Define the function to get pdq values for ARIMA model
def get_pdq_values():
    p = range(0, 6)
    d = range(0, 2)
    q = range(0, 2)
    pdq = [(x[0], x[1], x[2]) for x in list(itertools.product(p, d, q))]
    return pdq

def main():
    st.title("CRYPTON")

    # List of top cryptocurrencies by market cap
    response = requests.get("https://min-api.cryptocompare.com/data/top/mktcapfull?limit=50&tsym=USD")
    cryptocurrencies = [c["CoinInfo"]["Name"] for c in response.json()["Data"]]

    # User inputs
    selected_cryptocurrency = st.selectbox("Choose a cryptocurrency to analyze", cryptocurrencies)
    time_period_choice = st.selectbox("Choose a time period to analyze", ["24 hours", "7 days", "12 months"])
    algorithm_choice = st.selectbox("Choose an algorithm to use for price prediction", ["Linear Regression", "ARIMA"])

    if st.button("Predict Prices"):
        # Set the CryptoCompare API parameters based on user inputs
        if time_period_choice == "24 hours":
            name_time_period_choice = "24 hours"
            CRYPTOCOMPARE_API_PARAMS = {
                "fsym": selected_cryptocurrency,
                "tsym": "USD",
                "limit": 24,
                "aggregate": 1
            }
        elif time_period_choice == "7 days":
            name_time_period_choice = "7 days"
            CRYPTOCOMPARE_API_PARAMS = {
                "fsym": selected_cryptocurrency,
                "tsym": "USD",
                "limit": 168,
                "aggregate": 1
            }
        elif time_period_choice == "12 months":
            name_time_period_choice = "12 months"
            CRYPTOCOMPARE_API_PARAMS = {
                "fsym": selected_cryptocurrency,
                "tsym": "USD",
                "limit": 365,
                "aggregate": 1
            }

        # Get the historical price data
        response = requests.get(CRYPTOCOMPARE_API_ENDPOINT, params=CRYPTOCOMPARE_API_PARAMS)
        if not response.ok:
            st.error("Could not retrieve historical price data. Please try again.")
            return

        history_data = response.json()["Data"]["Data"]
        if not history_data:
            st.error("Historical price data is empty. Please try again.")
            return

        # Prepare the historical price data for machine learning
        timestamps = [datetime.fromtimestamp(h["time"]).timestamp() * 1000 for h in history_data]
        prices = [h["close"] for h in history_data]
        timestamps = [[t] for t in timestamps]

        # Train the machine learning model
        if algorithm_choice == "Linear Regression":
            model = LinearRegression()
            model.fit(timestamps, prices)
        elif algorithm_choice == "ARIMA":
            pdq = get_pdq_values()
            best_aic = float("inf")
            best_pdq = None
            for param in pdq:
                try:
                    model = ARIMA(prices, order=param)
                    result = model.fit()
                    if result.aic < best_aic:
                        best_aic = result.aic
                        best_pdq = param
                except:
                    continue
            model = ARIMA(prices, order=best_pdq)
            model_fit = model.fit()
        
        # Generate the price predictions
        current_time = datetime.now()
        predictions = []
        if time_period_choice == "24 hours":
            for i in range(24):
                next_time = current_time + timedelta(hours=i+1)
                next_timestamp = int(next_time.timestamp() * 1000)
                if algorithm_choice == "Linear Regression":
                    next_price = model.predict([[next_timestamp]])[0]
                elif algorithm_choice == "ARIMA":
                    next_price = model_fit.forecast(steps=1)[0]
                predictions.append((next_time, next_price))
        elif time_period_choice == "7 days":
            for i in range(7):
                next_time = current_time + timedelta(days=i+1)
                next_timestamp = int(next_time.timestamp() * 1000)
                if algorithm_choice == "Linear Regression":
                    next_price = model.predict([[next_timestamp]])[0]
                elif algorithm_choice == "ARIMA":
                    next_price = model_fit.forecast(steps=1)[0]
                predictions.append((next_time, next_price))
        elif time_period_choice == "12 months":
            for i in range(12):
                next_date_time = current_time + timedelta(days=30*(i+1))
                next_date = next_date_time.date()
                next_timestamp = int(next_date_time.timestamp() * 1000)
                if algorithm_choice == "Linear Regression":
                    next_price = model.predict([[next_timestamp]])[0]
                elif algorithm_choice == "ARIMA":
                    next_price = model_fit.forecast(steps=1)[0]
                predictions.append((next_date, next_price))

        # Display the predictions
        st.write(f"Price predictions for: {selected_cryptocurrency}")
        st.write(f"Algorithm: {algorithm_choice}")
        st.write(f"Time period: {name_time_period_choice}")

        st.table(predictions)

        # Save the predictions to an Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Price Predictions"
        ws.cell(1, 1, value="Date")
        ws.cell(1, 2, value="Price")
        for i, (date, price) in enumerate(predictions):
            ws.cell(i+2, 1, value=date)
            ws.cell(i+2, 2, value=price)
            ws.cell(i+2, 1).number_format = "mm/dd/yyyy hh:mm"

        date_time_string = datetime.now().strftime("%m%d%Y-%H%M%S")
        file_path = f"reports/{algorithm_choice}-{name_time_period_choice.replace(' ', '-')}-{selected_cryptocurrency}-{date_time_string}.xlsx"
        wb.save(file_path)
        st.success(f"Saved to: {file_path}")
        st.download_button(label="Download Predictions", data=open(file_path, 'rb').read(), file_name=os.path.basename(file_path))

if __name__ == "__main__":
    main()
